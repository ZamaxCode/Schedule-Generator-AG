import openpyxl
from openpyxl.workbook.workbook import Workbook
import materias as mat

def read_excel(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    lista_materias = []
    i = 2
    while True:
        if(ws.cell(row=i, column=1).value):
            materia = ws.cell(row=i, column=1).value
            seccion = ws.cell(row=i, column=2).value
            profesor = ws.cell(row=i, column=3).value
            dias_horas = ws.cell(row=i, column=4).value
            creditos = ws.cell(row=i, column=5).value
            calificacion = ws.cell(row=i, column=6).value
            m = mat.Materia(materia, seccion, profesor, dias_horas, creditos, calificacion)
            lista_materias.append(m)
            i+=1
        else:
            break
    return lista_materias

def write_excel(lista_materias):
    wb = Workbook()
    ws = wb.active
    
    ws.cell(row=1, column=1, value='Materia')
    ws.cell(row=1, column=2, value='Seccion')
    ws.cell(row=1, column=3, value='Profesor')
    ws.cell(row=1, column=4, value='Dias-Horas')
    ws.cell(row=1, column=5, value='Creditos')
    ws.cell(row=1, column=6, value='Calificacion')

    i=2
    for materia in lista_materias:
        ws.cell(row=i, column=1, value=materia._materia)
        ws.cell(row=i, column=2, value=materia._seccion)
        ws.cell(row=i, column=3, value=materia._profesor)
        ws.cell(row=i, column=4, value=materia._dias_horas)
        ws.cell(row=i, column=5, value=materia._creditos)
        ws.cell(row=i, column=6, value=materia._calificacion)
        i+=1

    wb.save("Horario_Generado.xlsx")
